# -*- coding: utf-8 -*-
"""
Created on Wed Aug  9 11:11:58 2017

@author: u1112166
"""

# -*- coding: utf-8 -*-
"""
Created on Tue Aug  8 10:42:07 2017

@author: KBEAN
"""


import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color


wb=openpyxl.Workbook()
ws=wb.active

wbstyle=openpyxl.load_workbook('Chubb North American Personal Lines_HUrev.xlsx')
wsStyle=wbstyle.get_sheet_by_name(name = 'Sheet1')


print(wsStyle['c11'].fill)


#==============================================================================
Header1 = openpyxl.styles.NamedStyle(name='Header1')
Header1.font = Font(name='Arial',color='FF002C77', sz=16.0)
wb.add_named_style(Header1)
#==============================================================================
tHeaderTeal=openpyxl.styles.NamedStyle('tHeaderTeal')
tHeaderTeal.font=Font(name='Arial',color='FFFFFFFF', sz=10.0,b=True)
tHeaderTeal.fill=PatternFill(fill_type='solid',start_color='FF00A8C8',end_color='FF00A8C8')
tHeaderTeal.border=Border(diagonalUp=False, diagonalDown=False, start=None, end=None, 
                   left=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'),
                   right=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'),
                   top=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'),
                   bottom=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'))
tHeaderTeal.alignment=Alignment(horizontal='center',vertical='center')
wb.add_named_style(tHeaderTeal)    
#==============================================================================
tHeaderBlue=openpyxl.styles.NamedStyle('tHeaderBlue')
tHeaderBlue.font=Font(name='Arial',color='FFFFFFFF', sz=10.0,b=True)
tHeaderBlue.fill=PatternFill(fill_type='solid',start_color='FF006D9E',end_color='FF006D9E')
tHeaderBlue.border=Border(diagonalUp=False, diagonalDown=False, start=None, end=None, 
                   left=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'),
                   right=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'),
                   top=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'),
                   bottom=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'))
tHeaderBlue.alignment=Alignment(horizontal='center',vertical='center')
wb.add_named_style(tHeaderBlue) 
#==============================================================================
tDataWhite=openpyxl.styles.NamedStyle('tDataWhite')
tDataWhite.font=Font(name='Arial',color='00000000', sz=10.0,b=False)
tDataWhite.fill=PatternFill(fill_type='solid',start_color='FFFFFFFF',end_color='FFFFFFFF')
tDataWhite.border=Border(diagonalUp=False, diagonalDown=False, start=None, end=None, 
                   left=Side(color='FF000000',border_style=None),
                   right=Side(color='FF000000',border_style=None),
                   top=Side(color='FF000000',border_style=None),
                   bottom=Side(color='FF000000',border_style=None))
tDataWhite.number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
tDataWhite.alignment=Alignment(horizontal='right',vertical='center')
wb.add_named_style(tDataWhite) 
#==============================================================================
tDataGray=openpyxl.styles.NamedStyle('tDataGray')
tDataGray.font=Font(name='Arial',color='00000000', sz=10.0,b=False)
tDataGray.fill=PatternFill(fill_type='solid',start_color='#EBEBEB',end_color='#EBEBEB')
tDataGray.border=Border(diagonalUp=False, diagonalDown=False, start=None, end=None, 
                   left=Side(color='FF000000',border_style=None),
                   right=Side(color='FF000000',border_style=None),
                   top=Side(color='FF000000',border_style=None),
                   bottom=Side(color='FF000000',border_style=None))
tDataGray.number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
tDataGray.alignment=Alignment(horizontal='right',vertical='center')
wb.add_named_style(tDataGray) 
#==============================================================================
ws['A1'].value='Chubb Ltd.'        
ws['A1'].style=Header1
ws['B2'].value='TIV'
ws['B2'].style=tHeaderTeal
ws['B3'].value='Chubb Legacy'
ws['B3'].style=tHeaderBlue
ws['B4'].value=25631
ws['B4'].style=tDataWhite
ws['B5'].value=2314646596
ws['B5'].style=tDataGray



for col in ws.columns:
    max_length = 0
    column = col[0].column # Get the column name
    for cell in col:
        if cell.coordinate in ws.merged_cells: # not check merge_cells
            continue
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column].width = adjusted_width
# ##print(Header1)
wb.save('example.xlsx')
 
 
#==============================================================================
