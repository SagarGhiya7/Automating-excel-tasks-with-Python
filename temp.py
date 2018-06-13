# Importing required packages
import pandas as pd
import numpy as np
import xlsxwriter
import os
import openpyxl
import pyodbc
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color


# Working Directory
os.chdir("C:/Users/u1112166/Desktop")
# Connecting to SQL Server
connection = pyodbc.connect('Driver={SQL Server};'

                                'Server=frkcat-rms4sql1;'

                                'Database=TRV_NECat_HU_2017_EDM15_forRe;')

                                
c = connection.cursor()
# 1st Sheet Data
fd = open('State_TIV1.sql', 'r')
sqlFile = fd.read()
fd.close()


c.execute(sqlFile)



qs = "select * from #StateTIV1"

c.execute(qs)
df = pd.DataFrame(c.fetchall(), columns = ['row'])


df1 = pd.read_sql(qs,connection)
# 1st sheet data

# 2nd sheet data
fd = open('State_TIV2.sql', 'r')
sqlFile1 = fd.read()
fd.close()

c.execute(sqlFile1)

qs1 = "select * from #StateTIV2"
c.execute(qs1)
df2 = pd.DataFrame(c.fetchall(), columns = ['row'])
df3 = pd.read_sql(qs1,connection)

df4 = df3.pivot_table(index = 'Admin1Code', values = 'Value', columns=['PORTNAME','State'] , aggfunc = np.sum, fill_value = 0)
#df4 = df4.reset_index(level=2, drop=True)

writer = pd.ExcelWriter('Automation.xlsx', engine='xlsxwriter')
#2nd sheet data

# Writing to excel
df1.to_excel(writer, sheet_name='Exposure by Geography', index = False, startrow = 7, startcol = 0)

df4.to_excel(writer, sheet_name='Pivot Table', startrow = 8, startcol=1)

workbook = writer.book
worksheet = writer.sheets['Exposure by Geography']
worksheet1 = writer.sheets['Pivot Table']
#format1 = workbook.add_format({'num_format': '#,##0'})

data_format1 = workbook.add_format({'num_format': '#,##0', 'font_name' : 'Arial', 'font_size': '10' })
#data_format2 = workbook.add_format({'num_format': '#,##0', 'font_name' : 'Arial', 'font_size':'10' })


#==============================================================================
for row in range(8,len(df1)+8):
        
    worksheet.set_row(row, cell_format=data_format1)
    
for row in range(8,len(df4)+11):    
    worksheet1.set_row(row , cell_format=data_format1)
   


worksheet.set_column('A:A', 16)
worksheet.set_column('B:B', 26)
worksheet.set_column('C:C', 14)
worksheet.set_column('D:D', 18)
worksheet.set_column('E:E', 18)
worksheet1.set_column('A:AO', 18)
link_format = workbook.add_format({'color': '#002C77', 'bold': True, 'size': 16, 'font_name': 'Arial'})
link1_format = workbook.add_format({'color': '#002c77', 'size': 14, 'font_name':'Arial'})
link2_format = workbook.add_format({'size':11, 'font_name' : 'Arial'})
link3_format = workbook.add_format({'color': '#FFFFFF','bg_color':'#00A8C8', 'bold' : True, 'size': 12, 'font_name':'Arial', 'align': 'center'})

worksheet.write('A1',  'Chubb Ltd.', link_format )
worksheet.write('A3', 'North American Personal Lines', link1_format)
worksheet.write('A5', 'Ground Up and Gross Exposed - Windstorm ', link2_format)
worksheet.write('A6', '$(000)', link2_format)
worksheet.merge_range('C7:E7', 'Chubb Legacy', link3_format)
worksheet1.write('A1',  'Chubb Ltd.', link_format )
worksheet1.write('A3', 'North American Personal Lines', link1_format)
worksheet1.write('A5', 'Ground Up and Gross Exposed - Windstorm ', link2_format)
worksheet1.write('A6', '$(000)', link2_format)
#worksheet1.merge_range('C7:E7', 'Chubb Legacy', link3_format)
#worksheet.write_column('A7', df1[,1])
writer.save()
workbook.close()





#Worksheet


wb=openpyxl.load_workbook('Automation.xlsx', read_only = False)
ws= wb.get_sheet_by_name('Exposure by Geography')
ws1= wb.get_sheet_by_name('Pivot Table')


Header1 = openpyxl.styles.NamedStyle(name='Header1')
Header1.font = Font(name='Arial',color='FF002C77', sz=16.0)
wb.add_named_style(Header1)

Header2 = openpyxl.styles.NamedStyle(name='Header2')
Header2.font = Font(name='Arial',color='FF002C77', sz=14.0)
wb.add_named_style(Header2)
#==============================================================================
tHeaderTeal=openpyxl.styles.NamedStyle('tHeaderTeal')
tHeaderTeal.font=Font(name='Arial',color='FFFFFFFF', sz=10.0,b=True)
tHeaderTeal.fill=PatternFill(fill_type='solid',start_color='FF00A8C8',end_color='FF00A8C8')
tHeaderTeal.border=Border(diagonalUp=False, diagonalDown=False, start=None, end=None, 
                   left=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'),
                   right=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'),
                   top=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'),
                   bottom=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'))
tHeaderTeal.alignment=Alignment(horizontal='center',vertical='center')
wb.add_named_style(tHeaderTeal)    
#==============================================================================
tHeaderBlue=openpyxl.styles.NamedStyle('tHeaderBlue')
tHeaderBlue.font=Font(name='Arial',color='FFFFFFFF', sz=10.0,b=True)
tHeaderBlue.fill=PatternFill(fill_type='solid',start_color='FF006D9E',end_color='FF006D9E')
tHeaderBlue.border=Border(diagonalUp=False, diagonalDown=False, start=None, end=None, 
                   left=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'),
                   right=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'),
                   top=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'),
                   bottom=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'))
tHeaderBlue.alignment=Alignment(horizontal='center',vertical='center')
wb.add_named_style(tHeaderBlue) 
#==============================================================================
tDataWhite=openpyxl.styles.NamedStyle('tDataWhite')
tDataWhite.font=Font(name='Arial',color='00000000', sz=10.0,b=False)
tDataWhite.fill=PatternFill(fill_type='solid',start_color='FFFFFFFF',end_color='FFFFFFFF')
#tDataWhite.border=Border(diagonalUp=False, diagonalDown=False, start=None, end=None, 
#                 left=Side(color='FF000000',border_style=None),
#                   bottom=Side(color='FF000000',border_style=None))
tDataWhite.number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
#tDataWhite.alignment=Alignment(horizontal='right',vertical='center', indent = 2.0)
wb.add_named_style(tDataWhite) 

tHeaderBlack=openpyxl.styles.NamedStyle('tHeaderBlack')
tHeaderBlack.font=Font(name='Arial', sz=11.0)
#tHeaderTeal.fill=PatternFill(fill_type='solid',start_color='FF00A8C8',end_color='FF00A8C8')
#tHeaderTeal.border=Border(diagonalUp=False, diagonalDown=False, start=None, end=None, 
#                   left=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'),
#                   right=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'),
#                   top=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'),
#                   bottom=Side(Color(rgb='bfbfbf',tint=-0.249946592608417),border_style='thin'))
#tHeaderTeal.alignment=Alignment(horizontal='center',vertical='center')
wb.add_named_style(tHeaderBlack)   
#==============================================================================
tDataGray=openpyxl.styles.NamedStyle('tDataGray')
tDataGray.font=Font(name='Arial',color='00000000', sz=10.0,b=False)
tDataGray.fill=PatternFill(fill_type='solid',start_color='FFEBEBEB',end_color='FFEBEBEB')
#tDataGray.border=Border(diagonalUp=False, diagonalDown=False, start=None, end=None, 
#                  right=Side(color='FF000000',border_style=None),
#                  top=Side(color='FF000000',border_style=None),
#                  bottom=Side(color='FF000000',border_style=None))
tDataGray.number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
#tDataGray.alignment=Alignment(horizontal='right',vertical='center', indent = 2.0)
wb.add_named_style(tDataGray) 

# Removing cell not needed
ws1['B11'].style= tDataWhite
ws1['B11'] = None

length1 = len(df1)
length2 = len(df4)
length_col1 = len(df1.columns)
length_col2 = len(df4.columns)

# Moving rows 1 step upwards in Pivot Table worksheet            

start_row = 12

start_col = 1

for row in ws1.iter_rows(min_row=start_row):
        for cell in row:
            ws1.cell(row = start_row-1, column = start_col, value=cell.value) 
            start_col += 1 
        start_row += 1 
        start_col = 1
# Removing additional last row

for col in ws1.iter_cols(min_row = start_row-1, max_row = start_row-1):
    for cell in col:
        cell.value = None
        cell.style = tDataWhite

# Alternate color styling for sheet1
for row in ws.rows:
    a = row[0].row  
    for cell in row:
        if cell.coordinate in ws.merged_cells:
            continue
        if len(str(cell.value)) > 0:
            if( a % 2 == 0):
                cell.style = tDataGray
            else:
                cell.style = tDataWhite
 
# Removing unfilled cells and making worksheet blank               
                
for col in ws.iter_cols(min_col=length_col1+1, max_col= 100):
     
    for cell in col:
       cell.style = tDataWhite      

for col in ws.iter_cols(min_col=1, max_col = length_col1, min_row = 1, max_row= 7):
     
    for cell in col:
       cell.style = tDataWhite  
       
for row in ws.iter_cols(min_row=length1+9, max_row= 800):
     
    for cell in row:
       cell.style = tDataWhite                  

# Alternate coloring for sheet2
for row in ws1.rows:
    b = row[0].row
   
    for cell in row:
             if cell.coordinate in ws.merged_cells:
                 continue
             
             if(b==9):
                 cell.style = tHeaderTeal
             if(b==10):
                 cell.style = tHeaderBlue
             if(b>10):
                
                 if len(str(cell.value)) > 0:
                     if( b % 2 == 0):
                             cell.style = tDataGray
                     else:
                             cell.style = tDataWhite

# Removing unfilled cells in sheet2                 
for col in ws1.iter_cols(min_col =1, max_col=1, min_row=7):
     
    for cell in col:
       cell.style = tDataWhite 
   
for col in ws1.iter_cols(min_col =length_col2+3, max_col=100):
     
    for cell in col:
       cell.style = tDataWhite 

for row in ws1.iter_cols(min_row = length2 + 11 , max_row=100):
    for cell in row:
       cell.style = tDataWhite 
for col in ws1.iter_cols(min_col = 1 , max_col = length_col2 + 2, min_row = 1, max_row=8):
    for cell in col:
       cell.style = tDataWhite 

# Individual cell styling              
ws['A1'].style=Header1
ws['A3'].style = Header2
ws['A5'].style = tHeaderBlack
ws['A6'].style = tHeaderBlack
ws['C7'].style = tHeaderTeal 
ws1['A1'].style= Header1    
ws1['A3'].style = Header2
ws1['A5'].style = tHeaderBlack
ws1['A6'].style = tHeaderBlack 
ws['A8'].style = tHeaderBlue
ws['B8'].style = tHeaderBlue
ws['C8'].style = tHeaderBlue
ws['D8'].style = tHeaderBlue
ws['E8'].style = tHeaderBlue
ws1['B9'].style = tDataWhite
ws1['B9'] = None
        

wb.save('Automation.xlsx')
wb.close()





                 
